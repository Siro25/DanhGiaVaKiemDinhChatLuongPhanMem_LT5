from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import User, Salary, WorkShift
from .forms import EmployeeCreateForm, EmployeeUpdateForm, LoginForm
from django.utils.http import url_has_allowed_host_and_scheme
from django.contrib import messages
from django.views.decorators.http import require_POST, require_http_methods
from functools import wraps
from django.db.models import Q
from customers.models import Customer

# Hàm kiểm tra vai trò
def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

def is_nhanvien(user):
    return user.is_authenticated and user.role == 'nhanvien'

def is_khachhang(user):
    return user.is_authenticated and user.role == 'khachhang'

# Decorator kiểm tra vai trò
def role_required(role):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if request.user.role != role:
                if request.user.role == 'admin':
                    return redirect('admin_dashboard')
                elif request.user.role == 'nhanvien':
                    return redirect('dashboard-nhanvien')
                else:
                    return redirect('dashboard_user')
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# Lightweight alias so `accounts.urls` can safely reference `user_login_view` at import time.
def user_login_view(request):
    """Dedicated login view for end-users. Uses `LoginForm` and renders
    `accounts/user_login.html`. Supports login by email or username and
    safe `next` redirects.
    """
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            ident = form.cleaned_data.get('identifier')
            password = form.cleaned_data.get('password')

            # Find username by email or username input
            username = None
            if ident:
                if '@' in ident:
                    u = User.objects.filter(email__iexact=ident).first()
                else:
                    u = User.objects.filter(username__iexact=ident).first()
                if u:
                    username = u.username

            user = authenticate(request, username=username, password=password) if username else None

            if user:
                # Block unapproved staff as in main login_view
                if getattr(user, 'role', None) == 'nhanvien' and getattr(user, 'status', 'approved') != 'approved':
                    if getattr(user, 'status', None) == 'pending':
                        messages.warning(request, 'Tài khoản nhân viên của bạn đang chờ được xác thực. Vui lòng liên hệ quản trị viên.')
                    else:
                        messages.error(request, 'Tài khoản nhân viên của bạn đã bị từ chối. Vui lòng liên hệ quản trị viên.')
                    return render(request, 'accounts/user_login.html', {'form': form})

                login(request, user)

                # Safe next redirect
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                    return redirect(next_url)

                # Redirect based on role
                role = getattr(user, 'role', None)
                if role == 'admin':
                    return redirect('admin_dashboard')
                elif role == 'nhanvien':
                    return redirect('dashboard-nhanvien')
                elif role == 'khachhang':
                    return redirect('dashboard-khachhang')
                else:
                    return redirect('dashboard')
            else:
                messages.error(request, 'Email hoặc mật khẩu không đúng')
    else:
        form = LoginForm()

    return render(request, 'accounts/user_login.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user_obj = None
        if email:
            user_obj = User.objects.filter(email=email).first()

        if user_obj is not None:
            user = authenticate(request, username=user_obj.username, password=password)
        else:
            user = None

        if user is not None:
            if user.role == 'nhanvien' and user.status != 'approved':
                if user.status == 'pending':
                    messages.warning(request, 'Tài khoản nhân viên của bạn đang chờ được xác thực. Vui lòng liên hệ quản trị viên.')
                else:
                    messages.error(request, 'Tài khoản nhân viên của bạn đã bị từ chối. Vui lòng liên hệ quản trị viên.')
                return render(request, 'accounts/login.html')
            
            login(request, user)
            # Chuyển hướng dựa trên vai trò
            if user.role == 'admin':
                return redirect('admin_dashboard')
            elif user.role == 'nhanvien':
                return redirect('dashboard-nhanvien')
            elif user.role == 'khachhang':
                return redirect('dashboard-khachhang')
            else:
                return redirect('dashboard')
        else:
            messages.error(request, 'Email hoặc mật khẩu không đúng')

    return render(request, 'accounts/login.html')

def admin_login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            error = 'Vui lòng nhập đầy đủ tên đăng nhập và mật khẩu'
            return render(request, 'accounts/admin_login.html', {'error': error})
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.role == 'admin':
                login(request, user)
                return redirect('admin_dashboard')
            else:
                error = 'Bạn không có quyền quản trị. Vui lòng đăng nhập bằng tài khoản admin.'
                return render(request, 'accounts/admin_login.html', {'error': error})
        else:
            error = 'Tên đăng nhập hoặc mật khẩu không đúng'
            return render(request, 'accounts/admin_login.html', {'error': error})
    
    return render(request, 'accounts/admin_login.html')

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Dashboard dành riêng cho admin."""
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Sum
    from parking.models import ParkingLot, ParkingSlot
    from customers.models import Customer
    from vehicles.models import Vehicle
    from cards.models import PaymentTransaction
    from parking.models import PricingSetting
    from decimal import Decimal
    
    # Thiết lập thời gian
    today = timezone.now().date()
    today_start = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    today_end = today_start + timedelta(days=1)
    
    # Thống kê cơ bản
    total_employees = User.objects.filter(role='nhanvien').count()
    total_customers = Customer.objects.filter(status='approved').count()  # Chỉ đếm khách đã phê duyệt
    total_parking_lots = ParkingLot.objects.count()
    total_parking_slots = ParkingSlot.objects.count()
    available_slots = ParkingSlot.objects.filter(status='available').count()
    occupied_slots = ParkingSlot.objects.filter(status='occupied').count()
    vehicles_parked = Vehicle.objects.filter(
        status='in',
        customer__status='approved'  # Chỉ đếm xe của khách đã được phê duyệt
    ).count()  # Xe đang gửi
    
    # TÍNH DOANH THU HÔM NAY (TẤT CẢ giao dịch đã thanh toán)
    today_revenue_total = PaymentTransaction.objects.filter(
        status='paid',
        created_at__gte=today_start,
        created_at__lt=today_end
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Format doanh thu để hiển thị đẹp
    if today_revenue_total >= 1000000:
        today_revenue_display = f"{today_revenue_total/1000000:.1f}M"
    elif today_revenue_total >= 1000:
        today_revenue_display = f"{today_revenue_total/1000:.0f}K"
    else:
        today_revenue_display = f"{today_revenue_total:.0f}"
    
    # Thống kê doanh thu theo tháng (6 tháng gần đây)
    monthly_revenue_stats = []
    current_month = today.month
    current_year = today.year
    
    for i in range(6):
        # Tính toán tháng và năm chính xác
        target_month = current_month - i
        target_year = current_year
        
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        
        # Tính doanh thu tháng đó
        month_start = timezone.make_aware(timezone.datetime.combine(
            today.replace(year=target_year, month=target_month, day=1), 
            timezone.datetime.min.time()
        ))
        
        if target_month == 12:
            next_month = 1
            next_year = target_year + 1
        else:
            next_month = target_month + 1
            next_year = target_year
        
        month_end = timezone.make_aware(timezone.datetime.combine(
            today.replace(year=next_year, month=next_month, day=1),
            timezone.datetime.min.time()
        ))
        
        # Tính TẤT CẢ doanh thu trong tháng đó (không phân biệt loại khách)
        month_total_revenue = PaymentTransaction.objects.filter(
            status='paid',
            created_at__gte=month_start,
            created_at__lt=month_end
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        monthly_revenue_stats.append({
            'month': f"{target_month:02d}/{target_year}",
            'revenue': float(month_total_revenue)  # Convert to float for JavaScript
        })
    
    monthly_revenue_stats.reverse()  # Sắp xếp từ cũ đến mới

    # Thống kê lượng xe theo ngày (7 ngày gần đây)
    from parking.models import ParkingRecord
    daily_vehicle_stats = []
    
    for i in range(6, -1, -1):  # 7 ngày, từ 6 ngày trước đến hôm nay
        target_date = today - timedelta(days=i)
        day_start = timezone.make_aware(timezone.datetime.combine(target_date, timezone.datetime.min.time()))
        day_end = day_start + timedelta(days=1)
        
        # Đếm số lượt xe vào trong ngày
        vehicle_count = ParkingRecord.objects.filter(
            entry_time__gte=day_start,
            entry_time__lt=day_end
        ).count()
        
        daily_vehicle_stats.append({
            'date': target_date.strftime('%d/%m/%Y'),
            'count': vehicle_count
        })

    context = {
        'total_employees': total_employees,
        'total_customers': total_customers,
        'total_parking_lots': total_parking_lots,
        'total_parking_slots': total_parking_slots,
        'available_slots': available_slots,
        'occupied_slots': occupied_slots,
        'vehicles_parked': vehicles_parked,
        'today_revenue': today_revenue_display,
        'today_revenue_raw': float(today_revenue_total),  # Giá trị thô để debug
        'monthly_revenue_stats': monthly_revenue_stats,  # Data cho biểu đồ doanh thu
        'daily_vehicle_stats': daily_vehicle_stats,  # Data cho biểu đồ lượng xe
    }
    return render(request, 'admin/dashboard_admin.html', context)

@login_required
@user_passes_test(is_nhanvien)
def nhanvien_dashboard(request):
    """Dashboard dành riêng cho nhân viên."""
    from django.utils import timezone
    from datetime import timedelta, date
    from django.db.models import Count, Q, Sum
    from vehicles.models import Vehicle
    from customers.models import Customer, MonthlySubscription
    from cards.models import PaymentTransaction
    from parking.models import PricingSetting, ParkingRecord
    from collections import defaultdict
    from decimal import Decimal
    import calendar
    
    today = timezone.now().date()
    
    # Thống kê xe ra/vào hôm nay - DÙNG PARKINGRECORD thay vì Vehicle
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    today_vehicles = ParkingRecord.objects.filter(
        entry_time__range=[today_start, today_end]
    ).count()
    
    # Thống kê khách gửi tháng CÒN HẠN (subscription chưa hết hạn)
    monthly_customers = MonthlySubscription.objects.filter(
        is_active=True,
        end_date__gte=today
    ).count()
    
    # Thống kê 7 ngày gần đây cho biểu đồ - TẤT CẢ các lượt gửi xe
    daily_stats = []
    daily_revenue_stats = []  # Thêm thống kê doanh thu theo ngày
    
    for i in range(7):
        day = today - timedelta(days=i)
        day_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i)
        day_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999) - timedelta(days=i)
        
        # Thống kê lượt gửi xe (ParkingRecord)
        vehicle_count = ParkingRecord.objects.filter(
            entry_time__range=[day_start, day_end]
        ).count()
        
        # Thống kê doanh thu ngày đó (TẤT CẢ giao dịch đã thanh toán)
        day_total_revenue = PaymentTransaction.objects.filter(
            status='paid',
            created_at__gte=day_start,
            created_at__lt=day_end
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        daily_stats.append({
            'date': day.strftime('%d/%m'),
            'vehicles': vehicle_count
        })
        
        daily_revenue_stats.append({
            'date': day.strftime('%d/%m'),
            'revenue': float(day_total_revenue)  # Convert to float for JavaScript
        })
    
    daily_stats.reverse()  # Sắp xếp từ cũ đến mới
    daily_revenue_stats.reverse()
    
    # Thống kê subscription mới theo 6 tháng gần đây
    current_month = today.month
    current_year = today.year
    
    monthly_stats = []
    for i in range(6):
        # Tính toán tháng và năm chính xác
        target_month = current_month - i
        target_year = current_year
        
        while target_month <= 0:
            target_month += 12
            target_year -= 1
            
        subscription_count = MonthlySubscription.objects.filter(
            start_date__month=target_month,
            start_date__year=target_year
        ).count()
        
        monthly_stats.append({
            'month': f"{target_month:02d}/{target_year}",
            'customers': subscription_count
        })
    monthly_stats.reverse()
    
    # Xe đang gửi hiện tại (ParkingRecord chưa có exit_time)
    current_vehicles = ParkingRecord.objects.filter(
        exit_time__isnull=True
    ).count()
    
    # Tổng thanh toán hôm nay
    today_payments = PaymentTransaction.objects.filter(
        created_at__range=[today_start, today_end],
        status='paid'
    ).count()
    
    # TÍNH TOÁN DOANH THU THỰC TẾ THEO THỜI GIAN
    
    # === DOANH THU HÔM NAY ===
    # Tính TẤT CẢ doanh thu đã thanh toán hôm nay (không phân biệt loại khách)
    today_revenue = PaymentTransaction.objects.filter(
        status='paid',
        created_at__range=[today_start, today_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # === DOANH THU TUẦN NÀY (7 ngày gần đây) ===
    week_start = today_start - timedelta(days=6)  # 7 ngày tính từ hôm nay
    week_revenue = PaymentTransaction.objects.filter(
        status='paid',
        created_at__range=[week_start, today_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # === DOANH THU THÁNG NÀY ===
    month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_revenue = PaymentTransaction.objects.filter(
        status='paid',
        created_at__range=[month_start, today_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    context = {
        'today_vehicles': today_vehicles,
        'monthly_customers': monthly_customers,
        'current_vehicles': current_vehicles,
        'today_payments': today_payments,
        'daily_stats': daily_stats,
        'daily_revenue_stats': daily_revenue_stats,  # Thêm dữ liệu doanh thu theo ngày
        'monthly_stats': monthly_stats,
        # Doanh thu theo thời gian
        'today_revenue': today_revenue,
        'week_revenue': week_revenue,
        'month_revenue': month_revenue,
    }
    
    return render(request, 'nhanvien/dashboard_nhanvien.html', context)

@login_required
@user_passes_test(is_admin)
def employee_management(request):
    """Quản lý nhân viên - chỉ dành cho admin"""
    # Lọc theo tìm kiếm
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    employees = User.objects.filter(role='nhanvien')
    
    if search:
        employees = employees.filter(
            Q(username__icontains=search) |
            Q(full_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    # Phân chia theo trạng thái
    pending_employees = employees.filter(status='pending')
    approved_employees = employees.filter(status='approved')
    rejected_employees = employees.filter(status='rejected')
    
    context = {
        'pending_employees': pending_employees,
        'approved_employees': approved_employees,
        'rejected_employees': rejected_employees,
        'all_employees': employees,
        'search': search,
        'status_filter': status_filter,
    }
    
    return render(request, 'admin/employee_management.html', context)

@login_required
@user_passes_test(is_admin)
def approve_employee(request, employee_id):
    """Phê duyệt tài khoản nhân viên"""
    if request.method == 'POST':
        employee = get_object_or_404(User, id=employee_id, role='nhanvien')
        employee.status = 'approved'
        employee.is_verified = True
        employee.save()
        messages.success(request, f'Tài khoản nhân viên {employee.username} đã được xác thực thành công.')
    
    return redirect('employee_management')

@login_required
@user_passes_test(is_admin)
def reject_employee(request, employee_id):
    """Từ chối tài khoản nhân viên"""
    if request.method == 'POST':
        employee = get_object_or_404(User, id=employee_id, role='nhanvien')
        employee.status = 'rejected'
        employee.is_verified = False
        employee.save()
        messages.success(request, f'Tài khoản nhân viên {employee.username} đã bị từ chối.')
    
    return redirect('employee_management')

@require_POST
def logout_view(request):
    """Logout only via POST to protect against CSRF/accidental GET logouts."""
    from django.contrib import messages
    
    # Clear all pending messages before logout
    storage = messages.get_messages(request)
    for message in storage:
        pass  # This consumes all messages
    storage.used = True
    
    logout(request)
    return redirect('login')

@login_required
@user_passes_test(is_nhanvien)
def nhanvien_settings(request):
    """Trang cài đặt cho nhân viên - hiển thị lương và rút lương"""
    from .models import Salary, SalaryWithdraw, WorkShift
    from django.utils import timezone
    
    # Tạo salary record nếu chưa có
    salary, created = Salary.objects.get_or_create(
        user=request.user,
        defaults={
            'basic_salary': 0,
        }
    )
    
    # Lấy lịch sử rút lương
    withdraws = SalaryWithdraw.objects.filter(user=request.user).order_by('-requested_at')[:10]
    
    # Lấy ca làm việc hiện tại (nếu có)
    current_shift = WorkShift.objects.filter(
        user=request.user, 
        status='working'
    ).first()
    
    # Xử lý form
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'withdraw':
            # Xử lý rút lương trực tiếp (không cần admin duyệt)
            amount = request.POST.get('amount')
            reason = request.POST.get('reason', '')
            
            if amount:
                try:
                    from decimal import Decimal
                    amount = Decimal(str(amount))
                    
                    # Logic mới: Kiểm tra lương cơ bản hiện tại (không cần tính số dư)
                    if amount > 0 and amount <= salary.basic_salary and salary.basic_salary > 0:
                        # Tạo record rút lương với status 'completed'
                        SalaryWithdraw.objects.create(
                            user=request.user,
                            amount=amount,
                            reason=reason,
                            status='completed'  # Tự động hoàn thành
                        )
                        
                        # Logic mới: Trừ trực tiếp từ lương cơ bản + cộng vào đã rút
                        salary.basic_salary -= amount  # Giảm lương cơ bản
                        salary.withdrawn += amount     # Tăng số đã rút (để theo dõi lịch sử)
                        salary.save()
                        
                        # Thông báo thành công
                        messages.success(request, 
                            f'✅ Đã rút {amount:,.0f} VND thành công!\n'
                            f'💰 Lương cơ bản còn lại: {salary.basic_salary:,.0f} VND'
                        )
                        return redirect('nhanvien_settings')
                    else:
                        if salary.basic_salary <= 0:
                            messages.error(request, '❌ Không có tiền để rút. Vui lòng chờ admin phát lương.')
                        else:
                            messages.error(request, f'❌ Số tiền không hợp lệ. Lương khả dụng: {salary.basic_salary:,.0f} VND.')
                except ValueError:
                    messages.error(request, '❌ Vui lòng nhập số tiền hợp lệ.')
        
        elif action == 'start_shift':
            # Bắt đầu ca làm việc
            if not current_shift:
                WorkShift.objects.create(
                    user=request.user,
                    start_time=timezone.now(),
                    status='working'
                )
                messages.success(request, 'Đã bắt đầu ca làm việc 5 giây (Demo mode)!')
                return redirect('nhanvien_settings')
            else:
                messages.warning(request, 'Bạn đang trong ca làm việc!')
        
        elif action == 'end_shift':
            # Kết thúc ca làm việc
            if current_shift:
                current_shift.end_time = timezone.now()
                current_shift.status = 'finished'
                current_shift.save()
                messages.success(request, 'Đã kết thúc ca làm việc!')
                return redirect('nhanvien_settings')
            else:
                messages.warning(request, 'Bạn không trong ca làm việc nào!')
    
    context = {
        'salary': salary,
        'withdraws': withdraws,
        'current_shift': current_shift,
    }
    
    return render(request, 'nhanvien/settings.html', context)

@login_required
def dashboard(request):
    """Chuyển hướng dựa trên vai trò"""
    user = request.user
    role = getattr(user, 'role', None)
    if role == 'admin':
        return redirect('admin_dashboard')
    elif role == 'nhanvien':
        return redirect('dashboard-nhanvien')
    elif role == 'khachhang':
        # Lấy dữ liệu cho khách hàng
        from cards.models import Notification
        from parking.models import ParkingRecord
        from vehicles.models import Vehicle
        
        # Lấy Customer profile
        try:
            customer = Customer.objects.get(user=user)
        except Customer.DoesNotExist:
            customer = None
        
        # Lấy thông báo chưa đọc
        notifications = Notification.objects.filter(user=user, read=False).order_by('-created_at')[:5]
        notifications_count = Notification.objects.filter(user=user, read=False).count()
        
        # Lấy thông tin xe
        vehicles = Vehicle.objects.filter(customer=customer) if customer else Vehicle.objects.none()
        
        # Kiểm tra xe đang trong bãi
        current_parking = None
        if customer:
            # Lấy record xe đang trong bãi (chưa có exit_time)
            current_parking = ParkingRecord.objects.filter(
                vehicle__customer=customer,
                exit_time__isnull=True
            ).select_related('vehicle', 'parking_lot').first()
        
        return render(request, 'customers/dashboard_user.html', {
            'notifications': notifications,
            'notifications_count': notifications_count,
            'vehicles': vehicles,
            'customer': customer,
            'current_parking': current_parking
        })

def signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email')
        full_name = request.POST.get('full_name')
        phone_number = request.POST.get('phone_number')
        address = request.POST.get('address', '')
        role = request.POST.get('role', 'khachhang')

        if not username or not password:
            messages.error(request, 'Vui lòng nhập đầy đủ tên đăng nhập và mật khẩu')
        elif password != confirm_password:
            messages.error(request, 'Mật khẩu xác nhận không khớp')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Tên đăng nhập đã tồn tại')
        elif email and User.objects.filter(email=email).exists():
            messages.error(request, 'Email đã được sử dụng')
        else:
            # Xác định trạng thái xác thực dựa trên vai trò
            is_verified = True if role == 'khachhang' else False
            status = 'approved' if role == 'khachhang' else 'pending'
            
            user = User.objects.create_user(
                username=username, 
                email=email, 
                password=password,
                role=role,
                full_name=full_name,
                phone_number=phone_number,
                is_verified=is_verified,
                status=status
            )
            
            # Tự động tạo Customer profile cho khách hàng
            if role == 'khachhang':
                Customer.objects.create(
                    user=user,
                    name=full_name or username,
                    phone=phone_number or '',
                    address=address or '',
                    customer_type='Khách vãng lai',  # Mặc định là khách vãng lai
                )
                messages.success(request, 'Tạo tài khoản thành công. Vui lòng đăng nhập.')
            elif role == 'nhanvien':
                messages.success(request, 'Tạo tài khoản nhân viên thành công. Vui lòng chờ quản trị viên xác thực trước khi đăng nhập.')
            else:
                messages.success(request, 'Tạo tài khoản thành công. Vui lòng đăng nhập.')
            
            return redirect('login')

    return render(request, 'accounts/signup.html')

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        # Check if user is admin
        return self.request.user.is_authenticated and self.request.user.role == 'admin'

class UserListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = User
    template_name = 'accounts/user_list.html'
    context_object_name = 'users'

class UserDetailView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = User
    template_name = 'accounts/user_detail.html'

class UserCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = User
    fields = ['username', 'email', 'first_name', 'last_name', 'role', 'phone_number', 'password']
    template_name = 'accounts/user_form.html'
    success_url = reverse_lazy('user-list')
    
    def form_valid(self, form):
        user = form.save(commit=False)
        password = form.cleaned_data.get('password')
        user.set_password(password)
        user.save()
        return super().form_valid(form)

class UserUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = User
    fields = ['username', 'email', 'first_name', 'last_name', 'role', 'phone_number']
    template_name = 'accounts/user_form.html'
    success_url = reverse_lazy('user-list')

class UserDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = User
    template_name = 'accounts/user_confirm_delete.html'
    success_url = reverse_lazy('user-list')

# Views quản lý nhân viên CRUD
@login_required
@user_passes_test(is_admin)
def employee_create(request):
    """Tạo tài khoản nhân viên mới"""
    if request.method == 'POST':
        form = EmployeeCreateForm(request.POST)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Tài khoản nhân viên {employee.username} đã được tạo thành công.')
            return redirect('employee_management')
    else:
        form = EmployeeCreateForm()
    
    return render(request, 'admin/employee_form.html', {
        'form': form,
        'title': 'Thêm tài khoản nhân viên mới',
        'action': 'create'
    })

@login_required
@user_passes_test(is_admin)
def employee_update(request, employee_id):
    """Cập nhật thông tin tài khoản nhân viên"""
    employee = get_object_or_404(User, id=employee_id, role='nhanvien')
    
    if request.method == 'POST':
        form = EmployeeUpdateForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f'Thông tin tài khoản nhân viên {employee.username} đã được cập nhật.')
            return redirect('employee_management')
    else:
        form = EmployeeUpdateForm(instance=employee)
    
    return render(request, 'admin/employee_form.html', {
        'form': form,
        'employee': employee,
        'title': f'Cập nhật thông tin - {employee.username}',
        'action': 'update'
    })

@login_required
@user_passes_test(is_admin)
def employee_detail(request, employee_id):
    """Xem chi tiết thông tin nhân viên"""
    employee = get_object_or_404(User, id=employee_id, role='nhanvien')
    
    return render(request, 'admin/employee_detail.html', {
        'employee': employee
    })

@login_required
@user_passes_test(is_admin)
def employee_delete(request, employee_id):
    """Xóa tài khoản nhân viên"""
    employee = get_object_or_404(User, id=employee_id, role='nhanvien')
    
    if request.method == 'POST':
        username = employee.username
        employee.delete()
        messages.success(request, f'Tài khoản nhân viên {username} đã được xóa.')
        return redirect('employee_management')
    
    return render(request, 'admin/employee_confirm_delete.html', {
        'employee': employee
    })


@user_passes_test(is_admin)
def reports_dashboard(request):
    """Trang báo cáo thống kê cho admin"""
    from django.db.models import Count, Sum
    from django.utils import timezone
    from datetime import timedelta
    from customers.models import Customer
    from vehicles.models import Vehicle
    from cards.models import PaymentTransaction
    
    # Thời gian
    now = timezone.now()
    today = now.date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    year_ago = today - timedelta(days=365)
    
    # Thống kê tổng quan (chỉ đếm khách hàng và phương tiện đã được phê duyệt)
    total_customers = Customer.objects.filter(is_active=True, status='approved').count()  # Chỉ đếm khách đã phê duyệt
    total_vehicles = Vehicle.objects.filter(customer__status='approved').count()  # Chỉ đếm xe của khách đã phê duyệt
    total_employees = User.objects.filter(role='nhanvien', is_verified=True).count()
    
    # Thống kê nhân viên chi tiết
    employees_active = User.objects.filter(role='nhanvien', is_active=True).count()
    employees_pending = User.objects.filter(role='nhanvien', status='pending').count()
    employees_approved = User.objects.filter(role='nhanvien', status='approved').count()
    employees_rejected = User.objects.filter(role='nhanvien', status='rejected').count()
    
    # Danh sách nhân viên với thông tin chi tiết
    employee_details = User.objects.filter(role='nhanvien').select_related().order_by('-date_joined')
    
    # Thống kê hoạt động của nhân viên (nếu có bảng log hoạt động)
    employee_stats = []
    for emp in employee_details:
        # Đếm số khách hàng được quản lý (nếu có trường người tạo)
        customers_managed = 0
        vehicles_managed = 0
        transactions_managed = 0
        
        try:
            # Nếu có trường created_by trong Customer
            if hasattr(Customer, 'created_by'):
                customers_managed = Customer.objects.filter(created_by=emp).count()
            
            # Nếu có trường created_by trong Vehicle  
            if hasattr(Vehicle, 'created_by'):
                vehicles_managed = Vehicle.objects.filter(created_by=emp).count()
            
            # Nếu có trường created_by trong PaymentTransaction
            if hasattr(PaymentTransaction, 'created_by'):
                transactions_managed = PaymentTransaction.objects.filter(created_by=emp).count()
        except:
            pass
            
        employee_stats.append({
            'employee': emp,
            'customers_managed': customers_managed,
            'vehicles_managed': vehicles_managed,
            'transactions_managed': transactions_managed,
            'total_activities': customers_managed + vehicles_managed + transactions_managed
        })
    
    # Sắp xếp theo tổng hoạt động
    employee_stats.sort(key=lambda x: x['total_activities'], reverse=True)
    
    # Thống kê phương tiện (chỉ tính xe của khách đã được phê duyệt)
    vehicles_in = Vehicle.objects.filter(
        status='in',
        customer__status='approved'  # Chỉ đếm xe của khách đã được phê duyệt
    ).count()
    vehicles_out = Vehicle.objects.filter(
        status='out',
        customer__status='approved'  # Chỉ đếm xe của khách đã được phê duyệt
    ).count()
    
    # Thống kê theo loại xe (tất cả xe trong hệ thống)
    vehicle_stats = Vehicle.objects.values('vehicle_type').annotate(count=Count('id')).order_by('-count')
    
    # Thống kê doanh thu theo thời gian
    try:
        from decimal import Decimal
        from datetime import datetime
        
        # Tính thời gian bắt đầu/kết thúc cho mỗi mốc thời gian (dùng timezone-aware datetime)
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end = today_start + timedelta(days=1)
        
        week_start = timezone.make_aware(datetime.combine(week_ago, datetime.min.time()))
        month_start = timezone.make_aware(datetime.combine(month_ago, datetime.min.time()))
        year_start = timezone.make_aware(datetime.combine(year_ago, datetime.min.time()))
        
        # Doanh thu hôm nay
        today_revenue = PaymentTransaction.objects.filter(
            created_at__gte=today_start,
            created_at__lt=today_end,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Doanh thu tuần này (7 ngày gần nhất)
        week_revenue = PaymentTransaction.objects.filter(
            created_at__gte=week_start,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Doanh thu tháng này (30 ngày gần nhất)
        month_revenue = PaymentTransaction.objects.filter(
            created_at__gte=month_start,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Doanh thu năm này (365 ngày gần nhất)
        year_revenue = PaymentTransaction.objects.filter(
            created_at__gte=year_start,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
    except Exception as e:
        print(f"Lỗi tính doanh thu: {e}")
        today_revenue = week_revenue = month_revenue = year_revenue = 0
    
    # Thống kê khách hàng mới (đã được phê duyệt)
    new_customers_today = Customer.objects.filter(
        created_at__date=today,
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()
    new_customers_week = Customer.objects.filter(
        created_at__date__gte=week_ago,
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()
    new_customers_month = Customer.objects.filter(
        created_at__date__gte=month_ago,
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()
    
    # Thống kê giao dịch
    try:
        total_transactions = PaymentTransaction.objects.count()
        paid_transactions = PaymentTransaction.objects.filter(status='paid').count()
        pending_transactions = PaymentTransaction.objects.filter(status='pending').count()
        cancelled_transactions = PaymentTransaction.objects.filter(status='cancelled').count()
    except:
        total_transactions = paid_transactions = pending_transactions = cancelled_transactions = 0
    
    # Thống kê khách hàng theo loại (tất cả khách hàng active)
    customer_type_stats = Customer.objects.filter(
        is_active=True
    ).values('customer_type').annotate(count=Count('id')).order_by('-count')
    
    # Top khách hàng (theo số lượng phương tiện, chỉ khách đã phê duyệt)
    top_customers = Customer.objects.filter(
        is_active=True,
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).annotate(
        vehicle_count=Count('vehicles')
    ).order_by('-vehicle_count')[:5]
    
    context = {
        # Tổng quan
        'total_customers': total_customers,
        'total_vehicles': total_vehicles,
        'total_employees': total_employees,
        
        # Thống kê nhân viên chi tiết
        'employees_active': employees_active,
        'employees_pending': employees_pending,
        'employees_approved': employees_approved,
        'employees_rejected': employees_rejected,
        'employee_stats': employee_stats[:10],  # Top 10 nhân viên hoạt động nhiều nhất
        'all_employees': employee_details,
        
        # Phương tiện
        'vehicles_in': vehicles_in,
        'vehicles_out': vehicles_out,
        'vehicle_stats': vehicle_stats,
        
        # Doanh thu
        'today_revenue': today_revenue,
        'week_revenue': week_revenue,
        'month_revenue': month_revenue,
        'year_revenue': year_revenue,
        
        # Khách hàng mới
        'new_customers_today': new_customers_today,
        'new_customers_week': new_customers_week,
        'new_customers_month': new_customers_month,
        
        # Giao dịch
        'total_transactions': total_transactions,
        'paid_transactions': paid_transactions,
        'pending_transactions': pending_transactions,
        'cancelled_transactions': cancelled_transactions,
        
        # Thống kê khách hàng theo loại
        'customer_type_stats': customer_type_stats,
        
        # Top khách hàng
        'top_customers': top_customers,
    }
    
    return render(request, 'admin/reports_dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def admin_customer_management(request):
    """Quản lý khách hàng dành riêng cho admin với chức năng nâng cao"""
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    from customers.models import Customer, MonthlySubscription
    from vehicles.models import Vehicle
    from cards.models import PaymentTransaction
    
    # Lọc và tìm kiếm
    search = request.GET.get('search', '').strip()
    subscription_status = request.GET.get('subscription_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset - tất cả khách hàng
    customers_qs = Customer.objects.all().order_by('-id')
    
    # Filter theo tìm kiếm
    if search:
        customers_qs = customers_qs.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(address__icontains=search) |
            Q(license_plate__icontains=search)
        )
    
    # Filter theo ngày tạo
    if date_from:
        try:
            from datetime import datetime
            from django.utils import timezone as tz
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            # Start of day
            date_from_start = tz.make_aware(datetime.combine(date_from_obj, datetime.min.time()))
            customers_qs = customers_qs.filter(created_at__gte=date_from_start)
        except Exception:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            from django.utils import timezone as tz
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            # End of day
            date_to_end = tz.make_aware(datetime.combine(date_to_obj, datetime.max.time()))
            customers_qs = customers_qs.filter(created_at__lte=date_to_end)
        except Exception:
            pass
    
    # Thống kê tổng quan (chỉ đếm khách hàng đã phê duyệt)
    total_customers = Customer.objects.filter(status='approved').count()  # Chỉ đếm khách đã phê duyệt
    monthly_customers = Customer.objects.filter(
        customer_type='Khách gửi tháng',
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()
    guest_customers = Customer.objects.filter(
        customer_type='Khách vãng lai',
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()
    
    # Khách hàng mới trong tuần (đã phê duyệt)
    week_ago = timezone.now().date() - timedelta(days=7)
    new_customers_week = Customer.objects.filter(
        created_at__date__gte=week_ago,
        status='approved'  # Chỉ đếm khách đã phê duyệt
    ).count()

    # Subscription sắp hết hạn (trong 7 ngày tới)
    next_week = timezone.now().date() + timedelta(days=7)
    expiring_subscriptions = MonthlySubscription.objects.filter(
        end_date__lte=next_week,
        end_date__gte=timezone.now().date(),
        is_active=True
    ).count()
    
    # Thống kê theo subscription status cho filter
    if subscription_status:
        today = timezone.now().date()
        if subscription_status == 'active':
            # Khách hàng có subscription còn hạn
            active_subscription_customers = MonthlySubscription.objects.filter(
                is_active=True,
                end_date__gte=today
            ).values_list('customer_id', flat=True)
            customers_qs = customers_qs.filter(id__in=active_subscription_customers)
        elif subscription_status == 'expired':
            # Khách hàng có subscription hết hạn
            expired_subscription_customers = MonthlySubscription.objects.filter(
                end_date__lt=today
            ).values_list('customer_id', flat=True)
            customers_qs = customers_qs.filter(id__in=expired_subscription_customers)
        elif subscription_status == 'none':
            # Khách hàng chưa có subscription
            customers_with_sub = MonthlySubscription.objects.values_list('customer_id', flat=True)
            customers_qs = customers_qs.exclude(id__in=customers_with_sub)
    
    # Annotate thêm thông tin
    customers_qs = customers_qs.annotate(
        vehicle_count=Count('vehicles'),
        total_paid=Sum('paymenttransaction__amount', filter=Q(paymenttransaction__status='paid'))
    )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(customers_qs, 20)  # 20 customers per page
    page_number = request.GET.get('page')
    customers = paginator.get_page(page_number)
    
    # Thêm thông tin subscription cho mỗi khách hàng
    today = timezone.now().date()
    from parking.models import PricingSetting
    from decimal import Decimal
    
    for customer in customers:
        # Lấy subscription mới nhất
        latest_sub = customer.subscriptions.order_by('-end_date').first()
        if latest_sub:
            customer.subscription_end = latest_sub.end_date
            customer.subscription_status = 'Còn hạn' if latest_sub.end_date >= today else 'Hết hạn'
            customer.days_until_expiry = (latest_sub.end_date - today).days if latest_sub.end_date >= today else 0
        else:
            customer.subscription_end = None
            customer.subscription_status = 'Chưa đăng ký'
            customer.days_until_expiry = None
        
        # Tổng tiền phải trả dựa trên loại xe đã đăng ký
        if customer.customer_type == 'Khách gửi tháng':
            # Tính dựa trên xe đã đăng ký
            total_monthly_amount = Decimal('0')
            customer_vehicles = customer.vehicles.all()
            
            for vehicle in customer_vehicles:
                try:
                    # Lấy giá tháng dựa trên loại xe
                    price = PricingSetting.get_price(vehicle.vehicle_type, 'monthly')
                    total_monthly_amount += Decimal(str(price))
                except Exception as e:
                    print(f"Lỗi lấy giá xe {vehicle.plate_number}: {e}")
            
            customer.total_paid_display = total_monthly_amount
        else:
            # Khách vãng lai: tính từ giao dịch thực tế
            customer.total_paid_display = customer.total_paid or Decimal('0')
    
    context = {
        'customers': customers,
        'total_customers': total_customers,
        'monthly_customers': monthly_customers,
        'guest_customers': guest_customers,
        'new_customers_week': new_customers_week,
        'expiring_subscriptions': expiring_subscriptions,
        
        # Form filters
        'search': search,
        'subscription_status': subscription_status,
        'date_from': date_from,
        'date_to': date_to,
        
        # Choices for filters
        'subscription_status_choices': [
            ('', 'Tất cả'),
            ('active', 'Còn hạn'),
            ('expired', 'Hết hạn'),
            ('none', 'Chưa đăng ký'),
        ],
    }
    
    return render(request, 'admin/customer_management.html', context)

@login_required 
@user_passes_test(is_admin)
def export_customers_excel(request):
    """Export danh sách khách hàng ra file Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime
    from customers.models import Customer, MonthlySubscription
    from django.db.models import Q
    
    # Lấy parameters từ request để export theo filter hiện tại
    search = request.GET.get('search', '').strip()
    subscription_status = request.GET.get('subscription_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply same filters as in admin_customer_management
    customers_qs = Customer.objects.all().order_by('-id')
    
    # Filter theo tìm kiếm
    if search:
        customers_qs = customers_qs.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(address__icontains=search) |
            Q(license_plate__icontains=search)
        )
    
    # Filter theo ngày tạo
    if date_from:
        try:
            from datetime import datetime
            from django.utils import timezone as tz
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            date_from_start = tz.make_aware(datetime.combine(date_from_obj, datetime.min.time()))
            customers_qs = customers_qs.filter(created_at__gte=date_from_start)
        except Exception:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            from django.utils import timezone as tz
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            date_to_end = tz.make_aware(datetime.combine(date_to_obj, datetime.max.time()))
            customers_qs = customers_qs.filter(created_at__lte=date_to_end)
        except Exception:
            pass
            
    # Filter theo subscription status
    if subscription_status:
        today = timezone.now().date()
        if subscription_status == 'active':
            active_subscription_customers = MonthlySubscription.objects.filter(
                is_active=True,
                end_date__gte=today
            ).values_list('customer_id', flat=True)
            customers_qs = customers_qs.filter(id__in=active_subscription_customers)
        elif subscription_status == 'expired':
            expired_subscription_customers = MonthlySubscription.objects.filter(
                end_date__lt=today
            ).values_list('customer_id', flat=True)
            customers_qs = customers_qs.filter(id__in=expired_subscription_customers)
        elif subscription_status == 'none':
            customers_with_sub = MonthlySubscription.objects.values_list('customer_id', flat=True)
            customers_qs = customers_qs.exclude(id__in=customers_with_sub)
    
    # Tạo workbook và worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách khách hàng"
    
    # Headers
    headers = [
        'STT', 'Tên khách hàng', 'Số điện thoại', 'Địa chỉ', 
        'Loại khách hàng', 'Loại xe', 'Biển số xe', 
        'Ngày tạo', 'Trạng thái subscription', 'Ngày hết hạn'
    ]
    
    # Style cho header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Thêm headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Thêm data
    today = timezone.now().date()
    for row_num, customer in enumerate(customers_qs, 2):
        # Lấy subscription info
        latest_sub = customer.subscriptions.order_by('-end_date').first()
        if latest_sub:
            subscription_status_text = 'Còn hạn' if latest_sub.end_date >= today else 'Hết hạn'
            subscription_end = latest_sub.end_date.strftime('%d/%m/%Y')
        else:
            subscription_status_text = 'Chưa đăng ký'
            subscription_end = ''
        
        data = [
            row_num - 1,  # STT
            customer.name,
            customer.phone,
            customer.address or '',
            customer.customer_type,
            customer.vehicle_type or '',
            customer.license_plate or '',
            customer.created_at.strftime('%d/%m/%Y %H:%M'),
            subscription_status_text,
            subscription_end
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Auto adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Tạo response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Tên file với timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'danh_sach_khach_hang_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Lưu workbook vào response
    wb.save(response)
    
    return response

@login_required 
@user_passes_test(is_admin)
def admin_payment_management(request):
    """Quản lý thu phí dành riêng cho admin với analytics nâng cao"""
    from django.db.models import Count, Sum, Q, Avg
    from django.utils import timezone
    from datetime import timedelta, datetime
    from cards.models import PaymentTransaction
    from vehicles.models import Vehicle
    from customers.models import Customer
    from parking.models import PricingSetting
    from decimal import Decimal
    import json
    
    # Filters
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('method', '')
    customer_type_filter = request.GET.get('customer_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    payments_qs = PaymentTransaction.objects.select_related('customer', 'vehicle').order_by('-created_at')
    
    # Filters
    if search:
        payments_qs = payments_qs.filter(
            Q(customer__name__icontains=search) |
            Q(customer__phone__icontains=search) |
            Q(vehicle__plate_number__icontains=search) |
            Q(reference__icontains=search)
        )
    
    if status_filter:
        payments_qs = payments_qs.filter(status=status_filter)
    
    if method_filter:
        payments_qs = payments_qs.filter(method=method_filter)
        
    if customer_type_filter:
        payments_qs = payments_qs.filter(customer__customer_type=customer_type_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(created_at__date__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(created_at__date__lte=date_to_obj)
        except:
            pass
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(payments_qs, 25)
    page_number = request.GET.get('page')
    payments = paginator.get_page(page_number)
    
    # ===== ANALYTICS & STATISTICS =====
    from django.utils import timezone
    now = timezone.now()
    today = now.date()
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # Tạo timezone-aware datetime ranges
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    yesterday_start = (timezone.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_end = (timezone.now() - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Doanh thu theo thời gian từ tất cả PaymentTransaction (không cần phân chia)
    today_revenue_transactions = PaymentTransaction.objects.filter(
        status='paid',
        created_at__range=[today_start, today_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    yesterday_revenue_transactions = PaymentTransaction.objects.filter(
        status='paid',
        created_at__range=[yesterday_start, yesterday_end]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    week_revenue_transactions = PaymentTransaction.objects.filter(
        status='paid',
        created_at__gte=timezone.now() - timedelta(days=7)
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    month_revenue_transactions = PaymentTransaction.objects.filter(
        status='paid',
        created_at__gte=timezone.now() - timedelta(days=30)
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Sử dụng doanh thu trực tiếp từ PaymentTransaction (không cộng thêm gì)
    try:
        # Doanh thu chi tiết theo loại khách hàng để hiển thị
        today_monthly_revenue = PaymentTransaction.objects.filter(
            customer__customer_type='Khách gửi tháng',
            status='paid',
            created_at__range=[today_start, today_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        today_guest_revenue = PaymentTransaction.objects.filter(
            customer__customer_type='Khách vãng lai',
            status='paid',
            created_at__range=[today_start, today_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Tương tự cho tuần
        week_start = today_start - timedelta(days=6)
        week_monthly_revenue = PaymentTransaction.objects.filter(
            customer__customer_type='Khách gửi tháng',
            status='paid',
            created_at__range=[week_start, today_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
    except Exception as e:
        print(f"Lỗi tính doanh thu khách gửi tháng: {e}")
        today_monthly_revenue = week_monthly_revenue = Decimal('0')
        today_guest_revenue = Decimal('0')
    
    # Sử dụng trực tiếp doanh thu từ PaymentTransaction (đã bao gồm tất cả)
    today_total_revenue = today_revenue_transactions  # Không cộng thêm
    week_total_revenue = week_revenue_transactions    # Không cộng thêm
    
    # Tính doanh thu tháng (30 ngày) từ PaymentTransaction thay vì Vehicle
    month_start = today_start - timedelta(days=29)  # 30 ngày tính từ hôm nay
    month_monthly_revenue = Decimal('0')
    try:
        month_monthly_revenue = PaymentTransaction.objects.filter(
            customer__customer_type='Khách gửi tháng',
            status='paid',  # Chỉ tính transaction đã approved
            created_at__range=[month_start, today_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    except Exception as e:
        print(f"Lỗi tính doanh thu khách gửi tháng cho tháng: {e}")
        month_monthly_revenue = Decimal('0')
    
    # Sử dụng trực tiếp doanh thu từ PaymentTransaction (đã bao gồm tất cả)
    month_total_revenue = month_revenue_transactions  # Không cộng thêm
    
    # Thống kê giao dịch
    total_transactions = PaymentTransaction.objects.count()
    paid_transactions = PaymentTransaction.objects.filter(status='paid').count()
    pending_transactions = PaymentTransaction.objects.filter(status='pending').count()
    cancelled_transactions = PaymentTransaction.objects.filter(status='cancelled').count()
    
    # Thống kê theo phương thức thanh toán
    payment_method_stats = PaymentTransaction.objects.filter(status='paid').values('method').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Thống kê theo loại khách hàng
    customer_type_revenue = PaymentTransaction.objects.filter(status='paid').values('customer__customer_type').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Top khách hàng theo doanh thu
    top_customers_revenue = Customer.objects.annotate(
        total_paid=Sum('paymenttransaction__amount', filter=Q(paymenttransaction__status='paid')),
        transaction_count=Count('paymenttransaction', filter=Q(paymenttransaction__status='paid'))
    ).filter(total_paid__isnull=False).order_by('-total_paid')[:10]
    
    # Xu hướng doanh thu 7 ngày gần đây
    daily_revenue_trend = []
    for i in range(7):
        date = today - timedelta(days=i)
        
        # Tạo start và end datetime cho ngày đó - dùng UTC timezone nhất quán
        start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i)
        end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999) - timedelta(days=i)
        
        # Doanh thu từ PaymentTransaction (bao gồm cả khách gửi tháng và vãng lai)
        daily_payment_revenue = PaymentTransaction.objects.filter(
            status='paid',  # Chỉ tính transaction đã approved
            created_at__range=[start_datetime, end_datetime]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Tổng doanh thu ngày đó (chỉ từ transaction đã approved)
        total_daily_revenue = daily_payment_revenue
        
        daily_revenue_trend.append({
            'date': date.strftime('%d/%m'),
            'revenue': float(total_daily_revenue)
        })
    
    daily_revenue_trend.reverse()  # Từ cũ đến mới
    
    # Growth rate
    try:
        if yesterday_revenue_transactions > 0:
            daily_growth = ((today_revenue_transactions - yesterday_revenue_transactions) / yesterday_revenue_transactions) * 100
        else:
            daily_growth = 100 if today_revenue_transactions > 0 else 0
    except:
        daily_growth = 0
    
    # Format numbers for display
    def format_currency(amount):
        if amount >= 1000000:
            return f"{amount/1000000:.1f}M"
        elif amount >= 1000:
            return f"{amount/1000:.0f}K"
        else:
            return f"{amount:.0f}"
    
    context = {
        'payments': payments,
        
        # Revenue Analytics
        'today_revenue': format_currency(today_total_revenue),
        'today_revenue_raw': float(today_total_revenue),
        'yesterday_revenue': format_currency(yesterday_revenue_transactions),
        'week_revenue': format_currency(week_total_revenue),
        'month_revenue': format_currency(month_total_revenue),
        'daily_growth': daily_growth,
        
        # Transaction Statistics  
        'total_transactions': total_transactions,
        'paid_transactions': paid_transactions,
        'pending_transactions': pending_transactions,
        'cancelled_transactions': cancelled_transactions,
        'success_rate': (paid_transactions / total_transactions * 100) if total_transactions > 0 else 0,
        
        # Analytics Data
        'payment_method_stats': payment_method_stats,
        'customer_type_revenue': customer_type_revenue,
        'top_customers_revenue': top_customers_revenue,
        'daily_revenue_trend': json.dumps(daily_revenue_trend),  # Serialize to JSON string
        
        # Filters
        'search': search,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'customer_type_filter': customer_type_filter,
        'date_from': date_from,
        'date_to': date_to,
        
        # Choices
        'status_choices': [
            ('', 'Tất cả trạng thái'),
            ('pending', 'Chờ thanh toán'),
            ('paid', 'Đã thanh toán'),
            ('cancelled', 'Đã hủy'),
        ],
        'method_choices': [
            ('', 'Tất cả phương thức'),
            ('cash', 'Tiền mặt'),
            ('card', 'Thẻ'),
            ('transfer', 'Chuyển khoản'),
        ],
        'customer_type_choices': [
            ('', 'Tất cả loại KH'),
            ('Khách gửi tháng', 'Khách gửi tháng'),
            ('Khách vãng lai', 'Khách vãng lai'),
        ],
    }
    
    return render(request, 'admin/payment_management.html', context)

@login_required 
@user_passes_test(is_admin)
def export_payments_excel(request):
    """Export danh sách giao dịch thanh toán ra file Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime
    from cards.models import PaymentTransaction
    from django.db.models import Q
    
    # Get filter parameters - same logic as admin_payment_management
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('method', '')
    customer_type_filter = request.GET.get('customer_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply same filters as in admin_payment_management
    payments_qs = PaymentTransaction.objects.select_related('customer', 'vehicle').order_by('-created_at')
    
    # Filters
    if search:
        payments_qs = payments_qs.filter(
            Q(customer__name__icontains=search) |
            Q(customer__phone__icontains=search) |
            Q(vehicle__plate_number__icontains=search) |
            Q(reference__icontains=search)
        )
    
    if status_filter:
        payments_qs = payments_qs.filter(status=status_filter)
    
    if method_filter:
        payments_qs = payments_qs.filter(method=method_filter)
        
    if customer_type_filter:
        payments_qs = payments_qs.filter(customer__customer_type=customer_type_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(created_at__date__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            payments_qs = payments_qs.filter(created_at__date__lte=date_to_obj)
        except:
            pass
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách giao dịch"
    
    # Headers
    headers = [
        'STT', 'Mã giao dịch', 'Khách hàng', 'SĐT', 'Biển số xe', 
        'Loại khách hàng', 'Số tiền', 'Phương thức', 'Trạng thái', 
        'Ngày tạo', 'Ghi chú'
    ]
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Add data
    for row_num, payment in enumerate(payments_qs, 2):
        # Status translation
        status_map = {
            'pending': 'Chờ thanh toán',
            'paid': 'Đã thanh toán', 
            'cancelled': 'Đã hủy'
        }
        
        # Method translation
        method_map = {
            'cash': 'Tiền mặt',
            'card': 'Thẻ',
            'transfer': 'Chuyển khoản'
        }
        
        data = [
            row_num - 1,  # STT
            payment.reference or f'PT{payment.id}',  # Transaction ID
            payment.customer.name if payment.customer else '',
            payment.customer.phone if payment.customer else '',
            payment.vehicle.plate_number if payment.vehicle else '',
            payment.customer.customer_type if payment.customer else '',
            float(payment.amount),
            method_map.get(payment.method, payment.method),
            status_map.get(payment.status, payment.status),
            payment.created_at.strftime('%d/%m/%Y %H:%M'),
            payment.reference or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            # Format currency column
            if col == 7:  # Amount column
                cell.number_format = '#,##0'
    
    # Auto adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'danh_sach_giao_dich_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def approve_transaction(request, transaction_id):
    """Duyệt giao dịch - chuyển từ 'pending' sang 'paid' và cập nhật customer status"""
    from django.http import JsonResponse
    from cards.models import PaymentTransaction
    from django.contrib import messages
    
    try:
        transaction = PaymentTransaction.objects.get(id=transaction_id)
        
        if transaction.status != 'pending':
            return JsonResponse({
                'success': False,
                'message': 'Giao dịch này không thể duyệt'
            })
        
        # Cập nhật status giao dịch
        transaction.status = 'paid'
        transaction.save()
        
        # Cập nhật trạng thái khách hàng thành "Đã duyệt"
        if transaction.customer:
            transaction.customer.status = 'approved'
            transaction.customer.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Giao dịch đã được xác nhận và cập nhật doanh thu thành công. Số tiền: {transaction.total:,} VNĐ',
            'new_status': 'paid',
            'amount': f'{transaction.total:,.0f} VNĐ'
        })
        
    except PaymentTransaction.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Không tìm thấy giao dịch'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Có lỗi xảy ra: {str(e)}'
        })

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def reject_transaction(request, transaction_id):
    """Từ chối giao dịch - chuyển từ 'pending' sang 'cancelled'"""
    from django.http import JsonResponse
    from cards.models import PaymentTransaction
    
    try:
        transaction = PaymentTransaction.objects.get(id=transaction_id)
        
        if transaction.status != 'pending':
            return JsonResponse({
                'success': False,
                'message': 'Giao dịch này không thể từ chối'
            })
        
        # Cập nhật status giao dịch
        transaction.status = 'cancelled'
        transaction.save()
        
        # Cập nhật trạng thái khách hàng về "Chờ duyệt" hoặc "Chưa đăng ký"
        if transaction.customer:
            transaction.customer.status = 'not_registered'
            transaction.customer.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Đã từ chối giao dịch',
            'new_status': 'cancelled'
        })
        
    except PaymentTransaction.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Không tìm thấy giao dịch'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Có lỗi xảy ra: {str(e)}'
        })

# ===========================================
# DATA CLEANING FUNCTIONS
# ===========================================

@login_required 
@user_passes_test(is_admin)
def clean_old_data(request):
    """Xóa dữ liệu cũ (khách hàng và giao dịch cũ hơn X ngày)"""
    from django.utils import timezone
    from datetime import timedelta
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect
    from cards.models import PaymentTransaction
    from customers.models import Customer
    from vehicles.models import Vehicle
    
    if request.method == 'POST':
        days = int(request.POST.get('days', 30))
        confirm = request.POST.get('confirm', False)
        
        if confirm:
            try:
                cutoff_date = timezone.now() - timedelta(days=days)
                
                with transaction.atomic():
                    # Xóa giao dịch cũ (chỉ xóa các giao dịch đã cancelled hoặc pending cũ)
                    old_transactions = PaymentTransaction.objects.filter(
                        created_at__lt=cutoff_date,
                        status__in=['cancelled', 'pending']  # Không xóa giao dịch đã thanh toán
                    )
                    transaction_count = old_transactions.count()
                    old_transactions.delete()
                    
                    # Xóa phương tiện của khách hàng không còn hoạt động
                    inactive_vehicles = Vehicle.objects.filter(
                        created_at__lt=cutoff_date,
                        customer__status='not_registered'  # Chỉ xóa xe của khách chưa đăng ký
                    )
                    vehicle_count = inactive_vehicles.count()
                    inactive_vehicles.delete()
                    
                    # Xóa khách hàng không hoạt động (chưa đăng ký và cũ)
                    inactive_customers = Customer.objects.filter(
                        created_at__lt=cutoff_date,
                        status='not_registered',  # Chỉ xóa khách chưa đăng ký
                        is_active=False
                    )
                    customer_count = inactive_customers.count()
                    inactive_customers.delete()
                    
                messages.success(request, 
                    f'Đã xóa thành công: {transaction_count} giao dịch cũ, '
                    f'{vehicle_count} phương tiện và {customer_count} khách hàng không hoạt động.')
                    
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra khi xóa dữ liệu: {str(e)}')
        else:
            messages.warning(request, 'Bạn phải xác nhận để thực hiện xóa dữ liệu.')
            
    return redirect('admin_payment_management')

@login_required 
@user_passes_test(is_admin)
def clean_cancelled_transactions(request):
    """Xóa tất cả giao dịch bị hủy"""
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect
    from cards.models import PaymentTransaction
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm', False)
        
        if confirm:
            try:
                with transaction.atomic():
                    cancelled_transactions = PaymentTransaction.objects.filter(status='cancelled')
                    count = cancelled_transactions.count()
                    cancelled_transactions.delete()
                    
                messages.success(request, f'Đã xóa thành công {count} giao dịch bị hủy.')
                    
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            messages.warning(request, 'Bạn phải xác nhận để thực hiện xóa dữ liệu.')
            
    return redirect('admin_payment_management')

@login_required 
@user_passes_test(is_admin)
def clean_duplicate_transactions(request):
    """Xóa giao dịch trùng lặp (cùng khách hàng, cùng phương tiện, cùng thời gian)"""
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.db.models import Count
    from cards.models import PaymentTransaction
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm', False)
        
        if confirm:
            try:
                with transaction.atomic():
                    # Tìm giao dịch trùng lặp
                    duplicates = PaymentTransaction.objects.values(
                        'customer', 'vehicle', 'created_at__date'
                    ).annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                    
                    deleted_count = 0
                    for dup in duplicates:
                        # Giữ lại giao dịch mới nhất, xóa các giao dịch cũ hơn
                        duplicate_transactions = PaymentTransaction.objects.filter(
                            customer=dup['customer'],
                            vehicle=dup['vehicle'],
                            created_at__date=dup['created_at__date']
                        ).order_by('-created_at')[1:]  # Bỏ qua giao dịch đầu tiên (mới nhất)
                        
                        for dt in duplicate_transactions:
                            dt.delete()
                            deleted_count += 1
                    
                messages.success(request, f'Đã xóa thành công {deleted_count} giao dịch trùng lặp.')
                    
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            messages.warning(request, 'Bạn phải xác nhận để thực hiện xóa dữ liệu.')
            
    return redirect('admin_payment_management')

@login_required 
@user_passes_test(is_admin)
def delete_specific_transaction(request, transaction_id):
    """Xóa giao dịch cụ thể theo ID"""
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.http import JsonResponse
    from cards.models import PaymentTransaction
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm', False)
        
        if confirm:
            try:
                transaction_obj = PaymentTransaction.objects.get(id=transaction_id)
                transaction_info = f"Mã GD: {transaction_obj.id}, Số tiền: {transaction_obj.amount}"
                transaction_obj.delete()
                
                messages.success(request, f'Đã xóa thành công giao dịch {transaction_info}')
                    
            except PaymentTransaction.DoesNotExist:
                messages.error(request, f'Không tìm thấy giao dịch với ID {transaction_id}')
            except Exception as e:
                messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        else:
            messages.warning(request, 'Bạn phải xác nhận để thực hiện xóa giao dịch.')
    else:
        # GET request - hiển thị confirmation
        try:
            transaction_obj = PaymentTransaction.objects.get(id=transaction_id)
            return JsonResponse({
                'transaction': {
                    'id': transaction_obj.id,
                    'amount': float(transaction_obj.amount),
                    'method': transaction_obj.method,
                    'status': transaction_obj.status,
                    'created_at': transaction_obj.created_at.strftime('%d/%m/%Y %H:%M'),
                    'customer': transaction_obj.customer.name if transaction_obj.customer else 'Không có'
                }
            })
        except PaymentTransaction.DoesNotExist:
            return JsonResponse({'error': f'Không tìm thấy giao dịch với ID {transaction_id}'})
            
    return redirect('admin_payment_management')


@role_required('admin')
def system_settings(request):
    """View để hiển thị trang cài đặt hệ thống"""
    # Get employees for dropdown
    employees = User.objects.filter(role='nhanvien')
    
    # Get salary data
    salary_data = []
    for emp in employees:
        salary, _ = Salary.objects.get_or_create(user=emp, defaults={'basic_salary': 0})
        
        # Chỉ đếm ca đã hoàn thành nhưng chưa được thanh toán
        completed_shifts = WorkShift.objects.filter(user=emp, status='finished').count()
        
        # Lương per shift mặc định nếu chưa có ca nào
        salary_per_shift = 200000  # Lương per shift cố định
        
        # Calculate total from unpaid completed shifts (chờ admin phát lương)
        pending_payment = salary_per_shift * completed_shifts
        
        # Update total_salary (chờ phát lương)
        salary.total_salary = pending_payment
        salary.save()
        
        # Check payment status
        status = 'Đã thanh toán' if completed_shifts == 0 else 'Chưa thanh toán'
        
        salary_data.append({
            'id': emp.id,
            'name': emp.full_name or emp.username,
            'salary_per_shift': salary_per_shift,
            'shifts_completed': completed_shifts,
            'total_salary': pending_payment,  # Tổng từ ca làm chờ phát
            'status': status
        })
    
    context = {
        'employees': employees,
        'salary_data': salary_data,
        'total_employees': employees.count()
    }
    
    return render(request, 'admin/system_settings.html', context)


@role_required('admin') 
def salary_management_api(request):
    """API để quản lý lương nhân viên"""
    from django.http import JsonResponse
    from .models import Salary, WorkShift, SalaryPayment
    from django.utils import timezone
    import json
    
    if request.method == 'GET':
        action = request.GET.get('action')
        
        if action == 'get_employees':
            # Lấy danh sách nhân viên
            employees = User.objects.filter(role='nhanvien')
            data = [{'id': emp.id, 'name': emp.full_name or emp.username} for emp in employees]
            return JsonResponse({'employees': data})
            
        elif action == 'get_salary_data':
            # Lấy thống kê lương nhân viên
            employees = User.objects.filter(role='nhanvien')
            salary_data = []
            
            for emp in employees:
                salary, _ = Salary.objects.get_or_create(user=emp, defaults={'basic_salary': 200000})
                completed_shifts = WorkShift.objects.filter(user=emp, status='finished').count()
                total_salary = salary.basic_salary * completed_shifts
                
                # Kiểm tra trạng thái thanh toán
                last_payment = SalaryPayment.objects.filter(employee=emp).order_by('-payment_date').first()
                status = 'Đã thanh toán' if last_payment and last_payment.shifts_count >= completed_shifts else 'Chưa thanh toán'
                
                salary_data.append({
                    'id': emp.id,
                    'name': emp.full_name or emp.username,
                    'salary_per_shift': salary.basic_salary,
                    'shifts_completed': completed_shifts,
                    'total_salary': total_salary,
                    'status': status
                })
            
            return JsonResponse({'salary_data': salary_data})
            
        elif action == 'get_employee_info':
            # Lấy thông tin chi tiết nhân viên
            employee_id = request.GET.get('employee_id')
            try:
                employee = User.objects.get(id=employee_id, role='nhanvien')
                salary, _ = Salary.objects.get_or_create(user=employee, defaults={'basic_salary': 200000})
                completed_shifts = WorkShift.objects.filter(user=employee, status='finished').count()
                total_salary = salary.basic_salary * completed_shifts
                
                last_payment = SalaryPayment.objects.filter(employee=employee).order_by('-payment_date').first()
                status = 'Đã thanh toán' if last_payment and last_payment.shifts_count >= completed_shifts else 'Chưa thanh toán'
                
                return JsonResponse({
                    'name': employee.full_name or employee.username,
                    'shifts': completed_shifts,
                    'salary': salary.basic_salary,
                    'total': total_salary,
                    'status': status
                })
            except User.DoesNotExist:
                return JsonResponse({'error': 'Nhân viên không tồn tại'})
                
        elif action == 'get_payment_history':
            # Lấy lịch sử thanh toán
            payments = SalaryPayment.objects.all().order_by('-payment_date')[:10]
            history_data = []
            
            for payment in payments:
                history_data.append({
                    'date': payment.payment_date.strftime('%Y-%m-%d'),
                    'employee': payment.employee.full_name or payment.employee.username,
                    'shifts': payment.shifts_count,
                    'amount': payment.amount,
                    'status': 'Hoàn thành'
                })
            
            return JsonResponse({'payment_history': history_data})
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')
        
        if action == 'update_salary':
            # Cập nhật lương nhân viên
            employee_id = data.get('employee_id')
            salary_per_shift = data.get('salary_per_shift')
            
            try:
                employee = User.objects.get(id=employee_id, role='nhanvien')
                salary, _ = Salary.objects.get_or_create(user=employee)
                salary.basic_salary = salary_per_shift
                salary.save()
                
                return JsonResponse({'success': True, 'message': 'Đã cập nhật lương thành công'})
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Nhân viên không tồn tại'})
                
        elif action == 'pay_salary':
            # Thanh toán lương - Logic mới: Admin nhập số tiền trực tiếp
            employee_id = data.get('employee_id')
            payment_amount = data.get('payment_amount', 200000)  # Số tiền admin muốn phát
            
            try:
                employee = User.objects.get(id=employee_id, role='nhanvien')
                salary, _ = Salary.objects.get_or_create(user=employee, defaults={'basic_salary': 0})
                
                # Logic mới: Admin phát số tiền cố định, không tính theo ca
                total_amount = float(payment_amount)  # Số tiền admin nhập
                
                # Convert to Decimal to match database field type
                from decimal import Decimal
                total_amount_decimal = Decimal(str(total_amount))
                
                # Debug: In thông tin chi tiết
                print(f"=== PHÁT LƯƠNG DEBUG ===")
                print(f"Nhân viên: {employee.username}")
                print(f"Số tiền admin phát: {total_amount:,}")
                print(f"Lương cơ bản trước khi phát: {salary.basic_salary:,}")
                
                # Tạo bản ghi thanh toán
                SalaryPayment.objects.create(
                    employee=employee,
                    amount=total_amount_decimal,  # Sử dụng Decimal
                    shifts_count=1,  # Ghi nhận 1 lần phát lương
                    payment_date=timezone.now(),
                    paid_by=request.user
                )
                
                # Đánh dấu các ca làm việc đã hoàn thành (nếu có)
                WorkShift.objects.filter(
                    user=employee, 
                    status='finished'
                ).update(status='paid')
                
                # Cập nhật lương: Gán số tiền admin phát thành basic_salary
                salary.basic_salary = total_amount_decimal  # Gán trực tiếp số tiền admin phát
                salary.total_salary = 0  # Reset total từ ca làm về 0
                salary.save()
                
                print(f"Lương cơ bản sau khi phát: {salary.basic_salary:,}")
                print(f"=========================")
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Đã phát {total_amount:,.0f} VNĐ cho {employee.full_name or employee.username}'
                })
                    
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Nhân viên không tồn tại'})
    
    return JsonResponse({'error': 'Method not allowed'})
