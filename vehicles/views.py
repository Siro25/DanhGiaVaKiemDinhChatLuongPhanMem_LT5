from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from .forms import VehicleForm, VehicleFilterForm
from .models import Vehicle
from django.core.paginator import Paginator
from customers.models import MonthlySubscription
from datetime import timedelta
from customers.models import Customer

def vehicle_list(request):
    form = VehicleFilterForm(request.GET or None)
    
    # Lấy tham số lọc loại khách hàng
    customer_type_filter = request.GET.get('customer_type', 'monthly')  # monthly, guest, all
    
    # Lọc theo loại khách hàng (sử dụng giá trị tiếng Anh trong database)
    if customer_type_filter == 'monthly':
        qs = Vehicle.objects.select_related('customer').filter(
            Q(customer__customer_type='monthly') | Q(customer__customer_type='Khách gửi tháng')
        ).order_by('-id')
        page_title = 'Quản lý phương tiện (Khách gửi tháng)'
    elif customer_type_filter == 'guest':
        qs = Vehicle.objects.select_related('customer').filter(
            Q(customer__customer_type='guest') | Q(customer__customer_type='Khách vãng lai')
        ).order_by('-id')
        page_title = 'Quản lý phương tiện (Khách vãng lai)'
    else:
        qs = Vehicle.objects.select_related('customer').all().order_by('-id')
        page_title = 'Quản lý tất cả phương tiện'

    q = request.GET.get('q', '').strip()
    vehicle_type = None
    owner = None
    date_from = None
    date_to = None
    per_page = 25

    if form.is_valid():
        q = form.cleaned_data.get('q', '').strip() or q
        vehicle_type = form.cleaned_data.get('vehicle_type')
        owner = form.cleaned_data.get('owner')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')

    if q:
        qs = qs.filter(
            Q(plate_number__icontains=q) |
            Q(customer__name__icontains=q)
        )
    if vehicle_type:
        qs = qs.filter(vehicle_type=vehicle_type)
    if owner:
        qs = qs.filter(customer__name__icontains=owner)
    if date_from:
        from django.utils import timezone as tz
        from datetime import datetime
        date_from_start = tz.make_aware(datetime.combine(date_from, datetime.min.time()))
        qs = qs.filter(check_in__gte=date_from_start)
    if date_to:
        from django.utils import timezone as tz
        from datetime import datetime
        date_to_end = tz.make_aware(datetime.combine(date_to, datetime.max.time()))
        qs = qs.filter(check_in__lte=date_to_end)

    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # For vehicles of monthly customers, compute subscription expiry/status
    today = timezone.now().date()
    for v in page_obj.object_list:
        v.subscription_end = None
        v.subscription_status = None
        if v.customer:
            sub = v.customer.subscriptions.order_by('-end_date').first()
            if sub:
                v.subscription_end = sub.end_date
                v.subscription_status = 'Còn hạn' if sub.end_date >= today else 'Hết hạn'

    querystring = request.GET.copy()
    if 'page' in querystring:
        querystring.pop('page')
    base_qs = querystring.urlencode()

    is_admin = getattr(request.user, 'role', None) == 'admin' or getattr(request.user, 'is_superuser', False)
    template = 'admin/vehicles/list.html' if is_admin else 'vehicles/list.html'
    
    # Thêm thống kê nhân viên cho admin
    employee_stats = []
    if is_admin:
        from accounts.models import User
        employees = User.objects.filter(role='nhanvien')
        for emp in employees:
            created_count = emp.created_vehicles.filter(
                Q(customer__customer_type='monthly') | Q(customer__customer_type='Khách gửi tháng')
            ).count()
            employee_stats.append({
                'employee': emp,
                'monthly_registrations': created_count
            })
    
    return render(request, template, {
        'form': form,
        'page_obj': page_obj,
        'base_qs': base_qs,
        'vehicles': page_obj,  # thêm để template dùng được
        'page_title': page_title,
        'customer_type_filter': customer_type_filter,
        'employee_stats': employee_stats if is_admin else None
    })

def vehicle_add(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            vehicle = form.save(commit=False)
            # Đảm bảo set check_in nếu chưa có
            if not vehicle.check_in:
                vehicle.check_in = timezone.now()
            if not vehicle.status:
                vehicle.status = 'in'
            # Mặc định là gói tháng vì chỉ cho khách gửi tháng
            if not vehicle.service_package:
                vehicle.service_package = 'monthly'
            # Ghi lại người tạo
            vehicle.created_by = request.user
            vehicle.save()
            
            # ===== QUY TRÌNH MỚI: Transaction sẽ được tạo tự động bởi post_save signal =====
            # Signal trong models.py sẽ tự động:
            # - Cập nhật customer status thành 'awaiting_approval' 
            # - Tạo PaymentTransaction với status 'pending'
            # - Tính toán giá dựa theo PricingSetting
            
            if vehicle.customer and vehicle.customer.customer_type == 'Khách gửi tháng':
                messages.success(request, f'✅ Đã thêm phương tiện. Giao dịch sẽ được tạo tự động.')
                messages.info(request, f'💰 Trạng thái: Chờ admin duyệt giao dịch')
            
            # Tự động tạo MonthlySubscription cho khách gửi tháng
            if vehicle.customer and vehicle.customer.customer_type == 'Khách gửi tháng':
                today = timezone.now().date()
                end_date = today + timedelta(days=30)  # Thêm 1 tháng
                
                # Kiểm tra xem khách hàng đã có subscription active chưa
                existing_sub = MonthlySubscription.objects.filter(
                    customer=vehicle.customer,
                    is_active=True,
                    end_date__gte=today
                ).first()
                
                if not existing_sub:
                    # Tạo subscription mới
                    subscription = MonthlySubscription.objects.create(
                        customer=vehicle.customer,
                        start_date=today,
                        end_date=end_date,
                        is_active=True
                    )
                    messages.info(request, f'📅 Đã tự động tạo đăng ký từ {today} đến {end_date}')
                else:
                    messages.info(request, f'📅 Khách hàng đã có đăng ký đến {existing_sub.end_date}')
            
            return redirect('vehicles:vehicle_list')  # hoặc 'vehicles:detail', vehicle.pk
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = VehicleForm()
    is_admin = getattr(request.user, 'role', None) == 'admin' or getattr(request.user, 'is_superuser', False)
    template = 'admin/vehicles/form.html' if is_admin else 'vehicles/form.html'
    return render(request, template, {'form': form})

def vehicle_edit(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật phương tiện.')
            return redirect('vehicles:vehicle_list')
        else:
            messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = VehicleForm(instance=vehicle)
    is_admin = getattr(request.user, 'role', None) == 'admin' or getattr(request.user, 'is_superuser', False)
    template = 'admin/vehicles/form.html' if is_admin else 'vehicles/form.html'
    return render(request, template, {'form': form, 'vehicle': vehicle})

def vehicle_detail(request, pk):
    v = get_object_or_404(Vehicle, pk=pk)
    is_admin = getattr(request.user, 'role', None) == 'admin' or getattr(request.user, 'is_superuser', False)
    template = 'admin/vehicles/detail.html' if is_admin else 'vehicles/detail.html'
    return render(request, template, {'vehicle': v})

    # Tính toán trạng thái đăng ký tháng
    today = timezone.now().date()
    v.subscription_end = None
    v.subscription_status = None
    v.subscription_days_left = None
    
    if v.customer and v.customer.customer_type == 'Khách gửi tháng':
        sub = v.customer.subscriptions.filter(is_active=True).order_by('-end_date').first()
        if sub:
            v.subscription_end = sub.end_date
            days_left = (sub.end_date - today).days
            v.subscription_days_left = days_left
            
            if days_left > 0:
                v.subscription_status = f'Còn hạn ({days_left} ngày)'
            elif days_left == 0:
                v.subscription_status = 'Hết hạn hôm nay'
            else:
                v.subscription_status = f'Hết hạn ({abs(days_left)} ngày)'
        else:
            v.subscription_status = 'Chưa đăng ký tháng'
    else:
        v.subscription_status = 'Khách vãng lai'
    
    return render(request, 'vehicles/detail.html', {'vehicle': v})

def vehicle_checkout(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if vehicle.status == 'in':
        vehicle.check_out = timezone.now()
        vehicle.status = 'out'
        vehicle.save()
        messages.success(request, 'Đã check-out.')
    return redirect('vehicles:vehicle_list')

def vehicle_delete(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    vehicle.delete()
    messages.success(request, 'Đã xóa phương tiện.')
    return redirect('vehicles:vehicle_list')

def export_vehicles_excel(request):
    """Export danh sách phương tiện ra file Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime
    from customers.models import MonthlySubscription
    from django.db.models import Q
    
    # Get filter parameters - same logic as vehicle_list
    q = request.GET.get('q', '').strip()
    vehicle_type = request.GET.get('vehicle_type', '')
    owner = request.GET.get('owner', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply same filters as in vehicle_list
    qs = Vehicle.objects.select_related('customer').filter(
        customer__customer_type='Khách gửi tháng'
    ).order_by('-id')
    
    if q:
        qs = qs.filter(
            Q(plate_number__icontains=q) |
            Q(customer__name__icontains=q)
        )
    if vehicle_type:
        qs = qs.filter(vehicle_type=vehicle_type)
    if owner:
        qs = qs.filter(customer__name__icontains=owner)
    if date_from:
        try:
            from django.utils import timezone as tz
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            date_from_start = tz.make_aware(datetime.combine(date_from_obj, datetime.min.time()))
            qs = qs.filter(check_in__gte=date_from_start)
        except Exception:
            pass
    if date_to:
        try:
            from django.utils import timezone as tz
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            date_to_end = tz.make_aware(datetime.combine(date_to_obj, datetime.max.time()))
            qs = qs.filter(check_in__lte=date_to_end)
        except Exception:
            pass
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách phương tiện"
    
    # Headers
    headers = [
        'STT', 'Biển số', 'Loại xe', 'Chủ xe', 'Ngày vào', 
        'Hết hạn đăng ký', 'Trạng thái đăng ký', 'Vị trí', 
        'Màu xe', 'Nhân viên đăng ký'
    ]
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Add data with subscription info
    today = timezone.now().date()
    for row_num, vehicle in enumerate(qs, 2):
        # Calculate subscription info
        subscription_end = None
        subscription_status = 'Chưa có thông tin'
        
        if vehicle.customer:
            latest_sub = vehicle.customer.subscriptions.order_by('-end_date').first()
            if latest_sub:
                subscription_end = latest_sub.end_date.strftime('%d/%m/%Y')
                days_left = (latest_sub.end_date - today).days
                if days_left > 0:
                    subscription_status = f'Còn hạn ({days_left} ngày)'
                elif days_left == 0:
                    subscription_status = 'Hết hạn hôm nay'
                else:
                    subscription_status = f'Hết hạn ({abs(days_left)} ngày)'
            else:
                subscription_status = 'Chưa đăng ký tháng'
        
        data = [
            row_num - 1,  # STT
            vehicle.plate_number,
            vehicle.get_vehicle_type_display(),
            vehicle.customer.name if vehicle.customer else '',
            vehicle.check_in.strftime('%d/%m/%Y %H:%M') if vehicle.check_in else '',
            subscription_end or '',
            subscription_status,
            vehicle.parking_slot or '',
            vehicle.color or '',
            vehicle.created_by.full_name if vehicle.created_by and vehicle.created_by.full_name else (vehicle.created_by.username if vehicle.created_by else '')
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Auto adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'danh_sach_phuong_tien_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response